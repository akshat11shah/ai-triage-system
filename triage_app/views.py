import json
import time
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseRedirect, JsonResponse
from django.urls import reverse
from django.contrib import messages
from .models import CustomerMessage, TriageResult, GroundTruth, KBArticle
from .triage_engine import TriageEngine
from .evaluator import seed_database, evaluate_system

def get_stats_dict():
    messages_list = CustomerMessage.objects.all().order_by('-created_at')
    triaged_count = TriageResult.objects.count()
    
    # Calculate aggregate stats
    p0_count = 0
    p1_count = 0
    p2_count = 0
    p3_count = 0
    needs_human_count = 0
    avg_confidence = 0.0
    avg_latency = 0.0
    total_cost = 0.0
    
    triage_results = TriageResult.objects.all()
    if triaged_count > 0:
        for r in triage_results:
            p = r.final_priority
            if p == 'P0': p0_count += 1
            elif p == 'P1': p1_count += 1
            elif p == 'P2': p2_count += 1
            elif p == 'P3': p3_count += 1
            
            if r.final_needs_human:
                needs_human_count += 1
                
            avg_confidence += r.confidence
            avg_latency += r.latency
            total_cost += r.cost
            
        avg_confidence /= triaged_count
        avg_latency /= triaged_count
        
    messages_data = []
    for msg in messages_list:
        triage = getattr(msg, 'triage', None)
        messages_data.append({
            'id': msg.id,
            'text': msg.text,
            'source': msg.source,
            'priority': triage.final_priority if triage else None,
            'needs_human': triage.final_needs_human if triage else None,
            'confidence': round(triage.confidence * 100, 1) if triage else None,
            'detail_url': reverse('message_detail', args=[msg.id])
        })
        
    return {
        'messages_list': messages_data,
        'triaged_count': triaged_count,
        'unprocessed_count': messages_list.count() - triaged_count,
        'p0_count': p0_count,
        'p1_count': p1_count,
        'p2_count': p2_count,
        'p3_count': p3_count,
        'needs_human_count': needs_human_count,
        'avg_confidence': round(avg_confidence * 100, 1),
        'avg_latency': round(avg_latency, 2),
        'total_cost': round(total_cost, 4),
    }

def dashboard(request):
    if request.GET.get('format') == 'json' or request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
        return JsonResponse(get_stats_dict())
        
    messages_list_all = CustomerMessage.objects.all().order_by('-created_at')
    
    # Pagination (10 messages per page)
    from django.core.paginator import Paginator
    paginator = Paginator(messages_list_all, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    triaged_count = TriageResult.objects.count()
    
    # Calculate aggregate stats
    p0_count = 0
    p1_count = 0
    p2_count = 0
    p3_count = 0
    needs_human_count = 0
    avg_confidence = 0.0
    avg_latency = 0.0
    total_cost = 0.0
    
    triage_results = TriageResult.objects.all()
    if triaged_count > 0:
        for r in triage_results:
            p = r.final_priority
            if p == 'P0': p0_count += 1
            elif p == 'P1': p1_count += 1
            elif p == 'P2': p2_count += 1
            elif p == 'P3': p3_count += 1
            
            if r.final_needs_human:
                needs_human_count += 1
                
            avg_confidence += r.confidence
            avg_latency += r.latency
            total_cost += r.cost
            
        avg_confidence /= triaged_count
        avg_latency /= triaged_count
        
    # Get evaluation metrics from session if run
    eval_metrics = request.session.get('eval_metrics', None)
    
    # Check if database is seeded
    db_is_seeded = CustomerMessage.objects.exists()
    kb_articles = KBArticle.objects.all().order_by('key')
    
    context = {
        'messages_list': page_obj,  # paginated page object
        'page_obj': page_obj,
        'total_messages_count': messages_list_all.count(),
        'triaged_count': triaged_count,
        'unprocessed_count': messages_list_all.count() - triaged_count,
        'p0_count': p0_count,
        'p1_count': p1_count,
        'p2_count': p2_count,
        'p3_count': p3_count,
        'needs_human_count': needs_human_count,
        'avg_confidence': round(avg_confidence * 100, 1),
        'avg_latency': round(avg_latency, 2),
        'total_cost': round(total_cost, 4),
        'eval_metrics': eval_metrics,
        'db_is_seeded': db_is_seeded,
        'kb_articles': kb_articles,
    }
    return render(request, 'triage_app/dashboard.html', context)

def message_detail(request, message_id):
    msg = CustomerMessage.objects.get(id=message_id)
    triage = getattr(msg, 'triage', None)
    
    # Calculate sequential Sr. No based on dashboard sort order (-created_at)
    sr_no = CustomerMessage.objects.filter(created_at__gt=msg.created_at).count() + 1
    
    if triage is None:
        TriageEngine._quota_exhausted = False
        engine = TriageEngine()
        triage_data = engine.triage_message(msg.text)
        triage = TriageResult.objects.create(
            message=msg,
            category=triage_data["category"],
            priority=triage_data["priority"],
            summary=triage_data["summary"],
            suggested_action=triage_data["suggested_action"],
            needs_human=triage_data["needs_human"],
            confidence=triage_data["confidence"],
            tool_calls_log=json.dumps(triage_data.get("tool_calls", [])),
            raw_json_response=triage_data.get("raw_json_response", ""),
            latency=triage_data["latency"],
            prompt_tokens=triage_data["prompt_tokens"],
            completion_tokens=triage_data["completion_tokens"],
            cost=triage_data["cost"]
        )
    
    tool_calls = []
    if triage and triage.tool_calls_log:
        try:
            tool_calls = json.loads(triage.tool_calls_log)
        except Exception:
            tool_calls = []
            
    formatted_json = ""
    if triage and triage.raw_json_response:
        try:
            parsed_json = json.loads(triage.raw_json_response)
            formatted_json = json.dumps(parsed_json, indent=2)
        except Exception:
            formatted_json = triage.raw_json_response
            
    context = {
        'message': msg,
        'sr_no': sr_no,
        'triage': triage,
        'tool_calls': tool_calls,
        'formatted_json': formatted_json,
        'priority_choices': TriageResult.PRIORITY_CHOICES,
    }
    return render(request, 'triage_app/detail.html', context)

def triage_single(request):
    if request.method == 'POST':
        text = request.POST.get('text', '').strip()
        source = request.POST.get('source', 'Web Dashboard')
        if text:
            msg = CustomerMessage.objects.create(text=text, source=source)
            TriageEngine._quota_exhausted = False
            engine = TriageEngine()
            triage_data = engine.triage_message(text)
            
            TriageResult.objects.create(
                message=msg,
                category=triage_data["category"],
                priority=triage_data["priority"],
                summary=triage_data["summary"],
                suggested_action=triage_data["suggested_action"],
                needs_human=triage_data["needs_human"],
                confidence=triage_data["confidence"],
                tool_calls_log=json.dumps(triage_data.get("tool_calls", [])),
                raw_json_response=triage_data.get("raw_json_response", ""),
                latency=triage_data["latency"],
                prompt_tokens=triage_data["prompt_tokens"],
                completion_tokens=triage_data["completion_tokens"],
                cost=triage_data["cost"]
            )
            
            if request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
                return JsonResponse({
                    'status': 'success',
                    'message': 'Message triaged successfully.',
                    'stats': get_stats_dict()
                })
                
            messages.success(request, f"Successfully triaged message #{msg.id}.")
            return redirect('message_detail', message_id=msg.id)
            
    return redirect('dashboard')

def triage_batch(request):
    un_triaged = CustomerMessage.objects.filter(triage__isnull=True)
    if un_triaged.exists():
        TriageEngine._quota_exhausted = False
        engine = TriageEngine()
        count = 0
        for msg in un_triaged:
            triage_data = engine.triage_message(msg.text)
            TriageResult.objects.create(
                message=msg,
                category=triage_data["category"],
                priority=triage_data["priority"],
                summary=triage_data["summary"],
                suggested_action=triage_data["suggested_action"],
                needs_human=triage_data["needs_human"],
                confidence=triage_data["confidence"],
                tool_calls_log=json.dumps(triage_data.get("tool_calls", [])),
                raw_json_response=triage_data.get("raw_json_response", ""),
                latency=triage_data["latency"],
                prompt_tokens=triage_data["prompt_tokens"],
                completion_tokens=triage_data["completion_tokens"],
                cost=triage_data["cost"]
            )
            count += 1
        
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
            return JsonResponse({
                'status': 'success',
                'message': f"Successfully triaged {count} pending messages.",
                'stats': get_stats_dict()
            })
            
        messages.success(request, f"Successfully triaged {count} pending messages.")
    else:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
            return JsonResponse({
                'status': 'success',
                'message': "No un-triaged messages found.",
                'stats': get_stats_dict()
            })
        messages.info(request, "No un-triaged messages found.")
        
    return redirect('dashboard')

def run_eval_view(request):
    TriageEngine._quota_exhausted = False
    metrics = evaluate_system()
    simple_metrics = {
        'num_cases': metrics['num_cases'],
        'category_accuracy': round(metrics['category_accuracy'], 1),
        'priority_accuracy': round(metrics['priority_accuracy'], 1),
        'needs_human_accuracy': round(metrics['needs_human_accuracy'], 1),
        'overall_agreement': round(metrics['overall_agreement'], 1),
        'avg_latency': round(metrics['avg_latency'], 2),
        'total_cost': round(metrics['total_cost'], 6),
        'avg_cost_per_msg': round(metrics['avg_cost_per_msg'], 6),
        'failures_count': len(metrics['failures']),
        'failures': metrics['failures'][:5]
    }
    request.session['eval_metrics'] = simple_metrics
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
        return JsonResponse({
            'status': 'success',
            'message': "Evaluation complete.",
            'eval_metrics': simple_metrics
        })
        
    messages.success(request, "Evaluation complete. Summary is displayed on the dashboard.")
    return redirect('dashboard')

def override_triage(request, result_id):
    triage = get_object_or_match(TriageResult, id=result_id)
    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'reset':
            triage.is_overridden = False
            triage.overridden_category = None
            triage.overridden_priority = None
            triage.overridden_needs_human = None
            triage.overridden_summary = None
            triage.overridden_suggested_action = None
            triage.save()
            msg_text = "Triage decision reset to AI original."
        else:
            category = request.POST.get('category', '').strip()
            priority = request.POST.get('priority', '')
            needs_human = request.POST.get('needs_human') == 'on'
            summary = request.POST.get('summary', '').strip()
            suggested_action = request.POST.get('suggested_action', '').strip()
            
            triage.is_overridden = True
            triage.overridden_category = category
            triage.overridden_priority = priority
            triage.overridden_needs_human = needs_human
            triage.overridden_summary = summary
            triage.overridden_suggested_action = suggested_action
            triage.save()
            msg_text = "Human override successfully saved."
            
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
            return JsonResponse({
                'status': 'success',
                'message': msg_text,
                'is_overridden': triage.is_overridden,
                'category': triage.final_category,
                'priority': triage.final_priority,
                'needs_human': triage.final_needs_human,
                'summary': triage.final_summary,
                'suggested_action': triage.final_suggested_action
            })
            
        messages.success(request, msg_text)
    return redirect('message_detail', message_id=triage.message.id)

def seed_data_view(request):
    TriageEngine._quota_exhausted = False
    msg_count, gt_count = seed_database()
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
        return JsonResponse({
            'status': 'success',
            'message': f"Seeding successful! Created {msg_count} messages and {gt_count} ground-truth test cases.",
            'stats': get_stats_dict()
        })
        
    messages.success(request, f"Seeding successful! Created {msg_count} messages and {gt_count} ground-truth test cases.")
    return redirect('dashboard')

def get_object_or_match(klass, *args, **kwargs):
    try:
        return klass.objects.get(*args, **kwargs)
    except klass.DoesNotExist:
        from django.http import Http404
        raise Http404("No message matches the given query.")

def manage_kb_article(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'delete':
            article_id = request.POST.get('id')
            article = get_object_or_404(KBArticle, id=article_id)
            article.delete()
            return JsonResponse({'status': 'success', 'message': 'Article deleted successfully.'})
            
        else:
            article_id = request.POST.get('id')
            key = request.POST.get('key', '').strip()
            content = request.POST.get('content', '').strip()
            
            if not key or not content:
                return JsonResponse({'status': 'error', 'message': 'Key and content are required.'}, status=400)
                
            if article_id:
                article = get_object_or_404(KBArticle, id=article_id)
                article.key = key
                article.content = content
                article.save()
                msg = "Article updated successfully."
            else:
                if KBArticle.objects.filter(key=key).exists():
                    return JsonResponse({'status': 'error', 'message': f"An article with the key '{key}' already exists."}, status=400)
                article = KBArticle.objects.create(key=key, content=content)
                msg = "Article created successfully."
                
            return JsonResponse({
                'status': 'success',
                'message': msg,
                'article': {
                    'id': article.id,
                    'key': article.key,
                    'content': article.content
                }
            })
            
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)

def import_excel(request):
    import openpyxl
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        # Verify file extension
        if not (excel_file.name.endswith('.xlsx') or excel_file.name.endswith('.xls')):
            if request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
                return JsonResponse({'status': 'error', 'message': 'Invalid file format. Please upload an Excel file (.xlsx or .xls).'}, status=400)
            messages.error(request, 'Invalid file format. Please upload an Excel file (.xlsx or .xls).')
            return redirect('dashboard')
            
        wb = None
        try:
            try:
                import io
                wb = openpyxl.load_workbook(io.BytesIO(excel_file.read()), data_only=True)
                sheet = wb.active
                
                # Find the message column by checking the first row for common headers
                headers = []
                for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)):
                    if cell:
                        headers.append(str(cell).strip().lower())
                    else:
                        headers.append('')
                
                target_col_idx = -1
                header_found = False
                possible_headers = ['message', 'messages', 'text', 'query', 'body', 'content', 'customer message', 'customer query']
                
                for header in possible_headers:
                    if header in headers:
                        target_col_idx = headers.index(header)
                        header_found = True
                        break
                        
                if not header_found:
                    # Check for substring match (e.g., 'query text', 'description')
                    for idx, h in enumerate(headers):
                        if any(p in h for p in possible_headers):
                            target_col_idx = idx
                            header_found = True
                            break
                            
                # If no header matches, default to the first column (index 0) and start from row 1
                if not header_found:
                    target_col_idx = 0
                    start_row = 1
                else:
                    start_row = 2
                    
                messages_created = []
                # Read rows
                for row in sheet.iter_rows(min_row=start_row, values_only=True):
                    if len(row) > target_col_idx:
                        val = row[target_col_idx]
                        if val:
                            msg_text = str(val).strip()
                            if msg_text:
                                # Create customer message in database
                                msg = CustomerMessage.objects.create(
                                    text=msg_text,
                                    source="Excel Import"
                                )
                                messages_created.append(msg)
            finally:
                if wb:
                    wb.close()
                            
            msg_text = f"Successfully imported {len(messages_created)} messages from Excel."
            if request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
                return JsonResponse({
                    'status': 'success',
                    'message': msg_text,
                    'stats': get_stats_dict()
                })
                
            messages.success(request, msg_text)
            
        except Exception as e:
            msg_text = f"Error: Failed to parse Excel file: {str(e)}"
            if request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
                return JsonResponse({'status': 'error', 'message': msg_text}, status=400)
            messages.error(request, msg_text)
            
        return redirect('dashboard')
        
    return JsonResponse({'status': 'error', 'message': 'No file uploaded.'}, status=400)

def generate_reply_api(request, message_id):
    if request.method == 'POST':
        try:
            msg = CustomerMessage.objects.get(id=message_id)
            triage = getattr(msg, 'triage', None)
            if not triage:
                return JsonResponse({'status': 'error', 'message': 'No triage result found.'}, status=400)
                
            engine = TriageEngine()
            
            # Extract KB articles if they were used
            kb_articles = ""
            if triage.tool_calls_log:
                try:
                    logs = json.loads(triage.tool_calls_log)
                    for call in logs:
                        if call.get('function') == 'search_knowledge_base' and call.get('result'):
                            kb_articles += call.get('result') + "\n"
                except:
                    pass
                    
            draft = engine.generate_draft_reply(
                message_text=msg.text,
                triage_category=triage.final_category,
                suggested_action=triage.suggested_action,
                kb_articles=kb_articles
            )
            
            triage.auto_reply_draft = draft
            triage.save()
            
            return JsonResponse({'status': 'success', 'draft': draft})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
            
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)
